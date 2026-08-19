import io
import re
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

from flask import current_app, jsonify, request, send_file
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from business_tools import (
    MAX_BULK_COUNT,
    MAX_BULK_FILE_BYTES,
    normalize_business_number,
    nts_status_request,
    serialize_status_item,
)

HEADER_KEYWORDS = {
    "사업자등록번호",
    "사업자번호",
    "사업자등록증번호",
    "사업자등록번호10자리",
}


def normalize_header(value) -> str:
    return re.sub(r"[^0-9A-Za-z가-힣]", "", str(value or "")).lower()


def cell_business_number(value) -> str:
    number = normalize_business_number(value)
    return number if len(number) == 10 else ""


def detect_business_column(sheet):
    """헤더 키워드를 우선 사용하고, 없으면 유효 번호가 가장 많이 있는 열을 보조 탐지한다."""
    max_scan_row = min(sheet.max_row, 10)
    scored = []

    for row_idx in range(1, max_scan_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            raw = sheet.cell(row_idx, col_idx).value
            header = normalize_header(raw)
            if not header:
                continue

            score = 0
            if header in HEADER_KEYWORDS:
                score = 100
            elif "사업자등록번호" in header:
                score = 95
            elif "사업자" in header and "번호" in header:
                score = 80

            if score:
                scored.append((score, row_idx, col_idx, str(raw)))

    if scored:
        scored.sort(reverse=True)
        best_score = scored[0][0]
        best = [item for item in scored if item[0] == best_score]
        unique_cols = {item[2] for item in best}
        if len(unique_cols) > 1:
            raise ValueError("사업자등록번호로 보이는 열이 여러 개입니다. 한 열의 제목을 '사업자등록번호'로 정리한 뒤 다시 시도해주세요.")
        _, header_row, col_idx, header_text = best[0]
        return header_row, col_idx, header_text

    # 헤더명이 일반적이지 않은 경우 1행을 헤더로 보고 데이터 패턴으로 보조 탐지
    column_counts = []
    for col_idx in range(1, sheet.max_column + 1):
        count = 0
        for row_idx in range(2, min(sheet.max_row, 202) + 1):
            if cell_business_number(sheet.cell(row_idx, col_idx).value):
                count += 1
        if count:
            column_counts.append((count, col_idx))

    if not column_counts:
        raise ValueError("사업자등록번호 열을 찾지 못했습니다. 첫 행의 열 이름을 '사업자등록번호'로 입력해주세요.")

    column_counts.sort(reverse=True)
    if len(column_counts) > 1 and column_counts[0][0] == column_counts[1][0]:
        raise ValueError("사업자등록번호 열을 자동으로 구분하기 어렵습니다. 첫 행의 열 이름을 '사업자등록번호'로 입력해주세요.")

    count, col_idx = column_counts[0]
    if count < 2:
        raise ValueError("사업자등록번호 열을 확실히 찾지 못했습니다. 첫 행의 열 이름을 '사업자등록번호'로 입력해주세요.")

    if cell_business_number(sheet.cell(1, col_idx).value):
        raise ValueError("첫 행이 데이터로 시작합니다. Excel 첫 행에 '사업자등록번호' 같은 열 제목을 추가해주세요.")

    return 1, col_idx, str(sheet.cell(1, col_idx).value or f"{get_column_letter(col_idx)}열")


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def format_date(value):
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 8:
        return f"{digits[:4]}.{digits[4:6]}.{digits[6:]}"
    return str(value or "")


def api_business_excel_enrich():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Excel(.xlsx) 파일을 선택해주세요."}), 400

    filename = file.filename
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"error": "원본 유지 점검은 XLSX 파일만 지원합니다."}), 400

    if request.content_length and request.content_length > MAX_BULK_FILE_BYTES:
        return jsonify({"error": "파일은 5MB 이하만 사용할 수 있습니다."}), 413

    raw = file.read(MAX_BULK_FILE_BYTES + 1)
    if len(raw) > MAX_BULK_FILE_BYTES:
        return jsonify({"error": "파일은 5MB 이하만 사용할 수 있습니다."}), 413

    try:
        workbook = load_workbook(io.BytesIO(raw), data_only=False)
        sheet = workbook.active
        header_row, business_col, detected_header = detect_business_column(sheet)

        row_numbers = []
        unique_numbers = []
        seen = set()

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            number = cell_business_number(sheet.cell(row_idx, business_col).value)
            if not number:
                continue
            row_numbers.append((row_idx, number))
            if number not in seen:
                seen.add(number)
                unique_numbers.append(number)

        if not unique_numbers:
            return jsonify({"error": "선택된 열에서 유효한 10자리 사업자등록번호를 찾지 못했습니다."}), 400
        if len(unique_numbers) > MAX_BULK_COUNT:
            return jsonify({"error": f"현재 버전은 한 파일에서 최대 {MAX_BULK_COUNT}개 사업자등록번호까지 점검할 수 있습니다. 파일을 나누어 다시 시도해주세요."}), 400

        items, error = nts_status_request(unique_numbers)
        if error:
            return jsonify({"error": error[0]}), error[1]

        item_map = {
            normalize_business_number(item.get("b_no")): item
            for item in items
            if normalize_business_number(item.get("b_no"))
        }
        result_map = {
            number: serialize_status_item(item_map.get(number, {}), number)
            for number in unique_numbers
        }

        start_col = sheet.max_column + 1
        headers = [
            "업무도구함_사업자상태",
            "업무도구함_과세유형",
            "업무도구함_폐업일",
            "업무도구함_확인필요",
            "업무도구함_조회일",
        ]

        header_source = sheet.cell(header_row, business_col)
        for offset, header in enumerate(headers):
            target = sheet.cell(header_row, start_col + offset, header)
            copy_cell_style(header_source, target)
            target.number_format = "General"

        checked_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        attention_count = 0

        for row_idx, number in row_numbers:
            result = result_map[number]
            status_name = result.get("statusName") or "확인 불가"
            attention = (not result.get("registered")) or ("계속" not in status_name)
            if attention:
                attention_count += 1

            values = [
                status_name,
                result.get("taxType") or "확인 불가",
                format_date(result.get("closureDate")),
                "확인 필요" if attention else "정상",
                checked_at,
            ]

            source = sheet.cell(row_idx, business_col)
            for offset, value in enumerate(values):
                target = sheet.cell(row_idx, start_col + offset, value)
                copy_cell_style(source, target)
                target.number_format = "General"

        widths = [20, 32, 14, 16, 22]
        for offset, width in enumerate(widths):
            letter = get_column_letter(start_col + offset)
            existing = sheet.column_dimensions[letter].width or 0
            sheet.column_dimensions[letter].width = max(existing, width)

        # 원본 시트·서식은 유지하고 결과 열만 오른쪽에 추가한다.
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        stem = Path(filename).stem
        download_name = f"{stem}_업무도구함_거래처점검.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_age=0,
        )
        response.headers["X-OTX-Checked-Count"] = str(len(unique_numbers))
        response.headers["X-OTX-Attention-Count"] = str(attention_count)
        response.headers["X-OTX-Detected-Header"] = detected_header.encode("utf-8", "ignore").hex()
        return response

    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        current_app.logger.exception("Business Excel enrich error: %s", exc.__class__.__name__)
        return jsonify({"error": "Excel 처리 중 오류가 발생했습니다. 파일 형식을 확인한 뒤 다시 시도해주세요."}), 500
