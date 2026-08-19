import csv
import io
import os
import re
from datetime import datetime, timezone
from urllib.parse import unquote

import requests
from flask import Blueprint, current_app, jsonify, render_template, request, send_file
from openpyxl import Workbook, load_workbook

business_bp = Blueprint("business", __name__)

NTS_STATUS_URL = "https://api.odcloud.kr/api/nts-businessman/v1/status"
BUSINESS_NUMBER_PATTERN = re.compile(r"(?<!\d)\d{3}-?\d{2}-?\d{5}(?!\d)")
MAX_BULK_COUNT = 100
MAX_BULK_FILE_BYTES = 5 * 1024 * 1024


def normalize_business_number(value: str) -> str:
    return re.sub(r"\D", "", str(value or ""))


def normalize_service_key(value: str) -> str:
    """Render에 Encoding/Decoding 키 중 어느 쪽이 저장돼도 requests에서 정상 전송되도록 정규화한다."""
    return unquote((value or "").strip())


def format_business_number(value: str) -> str:
    number = normalize_business_number(value)
    if len(number) != 10:
        return number
    return f"{number[:3]}-{number[3:5]}-{number[5:]}"


def extract_business_numbers(value) -> list[str]:
    text = str(value or "")
    return [normalize_business_number(match) for match in BUSINESS_NUMBER_PATTERN.findall(text)]


def unique_numbers(values, limit=MAX_BULK_COUNT + 1) -> list[str]:
    seen = set()
    result = []
    for value in values:
        number = normalize_business_number(value)
        if len(number) != 10 or number in seen:
            continue
        seen.add(number)
        result.append(number)
        if len(result) >= limit:
            break
    return result


def nts_status_request(numbers: list[str]):
    service_key = normalize_service_key(os.getenv("NTS_SERVICE_KEY", ""))
    if not service_key:
        return None, ("사업자 조회 서비스가 아직 설정되지 않았습니다.", 503)

    try:
        response = requests.post(
            NTS_STATUS_URL,
            params={"serviceKey": service_key, "returnType": "JSON"},
            json={"b_no": numbers},
            headers={"Accept": "application/json"},
            timeout=12,
        )
    except requests.Timeout:
        return None, ("국세청 조회 응답이 지연되고 있습니다. 잠시 후 다시 시도해주세요.", 504)
    except requests.RequestException as exc:
        current_app.logger.warning("NTS API connection error: %s", exc.__class__.__name__)
        return None, ("국세청 조회 서비스에 연결할 수 없습니다. 잠시 후 다시 시도해주세요.", 502)

    if not response.ok:
        provider_message = (response.text or "").strip().replace("\n", " ")[:500]
        current_app.logger.warning(
            "NTS API HTTP error status=%s body=%s",
            response.status_code,
            provider_message,
        )
        if response.status_code in (401, 403):
            return None, ("공공데이터포털 인증키가 아직 활성화되지 않았거나 유효하지 않습니다. 잠시 후 다시 시도해주세요.", 503)
        if response.status_code == 429:
            return None, ("오늘의 국세청 API 조회 한도에 도달했습니다. 잠시 후 다시 시도해주세요.", 503)
        if response.status_code in (400, 411, 413):
            return None, ("국세청 API 요청 형식을 확인할 수 없습니다. 관리자에게 문의해주세요.", 502)
        return None, ("국세청 조회 서비스가 일시적으로 응답하지 않습니다. 잠시 후 다시 시도해주세요.", 502)

    try:
        data = response.json()
    except ValueError:
        current_app.logger.warning("NTS API returned non-JSON response")
        return None, ("국세청 조회 결과를 해석할 수 없습니다. 잠시 후 다시 시도해주세요.", 502)

    return data.get("data") or [], None


def serialize_status_item(item: dict, fallback_number: str = "") -> dict:
    business_number = normalize_business_number(item.get("b_no") or fallback_number)
    status_code = (item.get("b_stt_cd") or "").strip()
    status_name = (item.get("b_stt") or "").strip()
    tax_type = (item.get("tax_type") or "").strip()
    compact_tax_type = tax_type.replace(" ", "")
    not_registered = "등록되지않은" in compact_tax_type
    registered = bool(status_code or status_name or tax_type) and not not_registered

    return {
        "businessNumber": format_business_number(business_number),
        "registered": registered,
        "statusCode": status_code,
        "statusName": status_name or ("확인 불가" if registered else "미등록 또는 확인 불가"),
        "taxType": tax_type or "확인 불가",
        "closureDate": (item.get("end_dt") or "").strip() or None,
        "taxTypeChangeDate": (item.get("tax_type_change_dt") or "").strip() or None,
        "invoiceApplyDate": (item.get("invoice_apply_dt") or "").strip() or None,
        "checkedAt": datetime.now(timezone.utc).isoformat(),
    }


@business_bp.get("/business-status")
def business_status_page():
    return render_template("business_status.html", page="business-status")


@business_bp.get("/business-bulk-status")
def business_bulk_status_page():
    return render_template("business_bulk_status.html", page="business-bulk-status")


@business_bp.post("/api/business/status")
def api_business_status():
    payload = request.get_json(silent=True) or {}
    business_number = normalize_business_number(payload.get("businessNumber", ""))

    if len(business_number) != 10:
        return jsonify({"error": "사업자등록번호 10자리를 입력해주세요."}), 400

    items, error = nts_status_request([business_number])
    if error:
        return jsonify({"error": error[0]}), error[1]
    if not items:
        return jsonify({"error": "조회 결과를 확인할 수 없습니다."}), 502

    return jsonify({"success": True, "data": serialize_status_item(items[0], business_number)})


@business_bp.post("/api/business/bulk-status")
def api_business_bulk_status():
    payload = request.get_json(silent=True) or {}
    values = payload.get("businessNumbers") or []
    if not isinstance(values, list):
        return jsonify({"error": "사업자등록번호 목록 형식이 올바르지 않습니다."}), 400

    numbers = unique_numbers(values)
    if not numbers:
        return jsonify({"error": "조회할 사업자등록번호를 입력해주세요."}), 400
    if len(numbers) > MAX_BULK_COUNT:
        return jsonify({"error": "한 번에 최대 100개까지 조회할 수 있습니다."}), 400

    items, error = nts_status_request(numbers)
    if error:
        return jsonify({"error": error[0]}), error[1]

    by_number = {
        normalize_business_number(item.get("b_no")): item
        for item in items
        if normalize_business_number(item.get("b_no"))
    }
    rows = [serialize_status_item(by_number.get(number, {}), number) for number in numbers]

    return jsonify(
        {
            "success": True,
            "data": rows,
            "meta": {
                "count": len(rows),
                "checkedAt": datetime.now(timezone.utc).isoformat(),
            },
        }
    )


@business_bp.post("/api/business/bulk-parse")
def api_business_bulk_parse():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "CSV 또는 XLSX 파일을 선택해주세요."}), 400
    if request.content_length and request.content_length > MAX_BULK_FILE_BYTES:
        return jsonify({"error": "파일은 5MB 이하만 사용할 수 있습니다."}), 413

    filename = file.filename.lower()
    raw = file.read(MAX_BULK_FILE_BYTES + 1)
    if len(raw) > MAX_BULK_FILE_BYTES:
        return jsonify({"error": "파일은 5MB 이하만 사용할 수 있습니다."}), 413

    found = []
    try:
        if filename.endswith(".csv"):
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("cp949")
            for row in csv.reader(io.StringIO(text)):
                for cell in row:
                    found.extend(extract_business_numbers(cell))
                    if len(unique_numbers(found)) > MAX_BULK_COUNT:
                        break
                if len(unique_numbers(found)) > MAX_BULK_COUNT:
                    break
        elif filename.endswith(".xlsx"):
            workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook.active
            stop = False
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    found.extend(extract_business_numbers(cell))
                    if len(unique_numbers(found)) > MAX_BULK_COUNT:
                        stop = True
                        break
                if stop:
                    break
            workbook.close()
        else:
            return jsonify({"error": "CSV 또는 XLSX 파일만 사용할 수 있습니다."}), 400
    except Exception as exc:
        current_app.logger.warning("Business bulk file parse error: %s", exc.__class__.__name__)
        return jsonify({"error": "파일을 읽을 수 없습니다. 파일 형식을 확인해주세요."}), 400

    numbers = unique_numbers(found)
    truncated = len(numbers) > MAX_BULK_COUNT
    numbers = numbers[:MAX_BULK_COUNT]
    if not numbers:
        return jsonify({"error": "파일에서 10자리 사업자등록번호를 찾지 못했습니다."}), 400

    return jsonify(
        {
            "success": True,
            "businessNumbers": [format_business_number(number) for number in numbers],
            "count": len(numbers),
            "truncated": truncated,
        }
    )


@business_bp.post("/api/business/bulk-export-xlsx")
def api_business_bulk_export_xlsx():
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "내보낼 조회 결과가 없습니다."}), 400
    if len(rows) > MAX_BULK_COUNT:
        return jsonify({"error": "한 번에 최대 100개의 결과만 내보낼 수 있습니다."}), 400

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "사업자 상태조회"
    headers = [
        "사업자등록번호",
        "등록여부",
        "사업자 상태",
        "과세유형",
        "폐업일자",
        "과세유형 전환일",
        "세금계산서 적용일",
        "조회시각",
    ]
    sheet.append(headers)

    for row in rows:
        if not isinstance(row, dict):
            continue
        sheet.append(
            [
                str(row.get("businessNumber") or ""),
                "등록" if row.get("registered") else "미등록/확인불가",
                str(row.get("statusName") or ""),
                str(row.get("taxType") or ""),
                str(row.get("closureDate") or ""),
                str(row.get("taxTypeChangeDate") or ""),
                str(row.get("invoiceApplyDate") or ""),
                str(row.get("checkedAt") or ""),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [18, 16, 18, 32, 14, 18, 20, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"business-status-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )
